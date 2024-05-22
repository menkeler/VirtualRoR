from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status,viewsets, filters
from .models import Category, ItemProfiling, ItemCopy, Inventory, ExportHistory
from transactions.models import Transaction,Inquiry,TransactionItem
from .serializers import CategorySerializer, ItemProfilingSerializer, ItemCopySerializer,ItemCopyCreateSerializer, InventorySerializer,InventoryCreateSerializer,ItemProfilingCreateSerializer,ExportHistorySerializer
from transactions.serializers import TransactionSerializer
from django.shortcuts import get_object_or_404
from rest_framework.decorators import action
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from datetime import datetime
from django.db.models import F, Value as V
from django.db.models.functions import Concat
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from django.http import JsonResponse
from django.db.models import Count
from django.db.models import Sum
from django.utils.dateparse import parse_date

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all().order_by('name') 
    serializer_class = CategorySerializer
    
class ItemProfilingPagination(PageNumberPagination):
    page_size = 30
    page_size_query_param = 'page_size'
    max_page_size = 1000
    
class ItemProfilingViewSet(viewsets.ModelViewSet):
    queryset = ItemProfiling.objects.all()
    pagination_class = ItemProfilingPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']
    
    # Different Serializer for Creation
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ItemProfilingCreateSerializer
        return ItemProfilingSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        category_param = self.request.query_params.get('category', None)
        returnable_param = self.request.query_params.get('returnable', None)

        if category_param:
            # Assuming 'category' is a field in your ItemProfiling model
            queryset = queryset.filter(category=category_param)

        if returnable_param is not None and returnable_param != '':
            # Assuming 'returnable' is a BooleanField in your ItemProfiling model
            queryset = queryset.filter(returnable=(returnable_param.lower() == 'true'))
            
        #alphabetical order
        queryset = queryset.order_by('name')

        return queryset
    

class ItemCopyViewSet(viewsets.ModelViewSet):
    queryset = ItemCopy.objects.all()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ItemCopyCreateSerializer
        return ItemCopySerializer

    def create(self, request, *args, **kwargs):
        print("Received data:", request.data)
        data_array = request.data if isinstance(request.data, list) else [request.data]
        response_data = []

        for data_item in data_array:
            profile_item_id = data_item.get('inventory')

            try:
                inventory = Inventory.objects.get(item=profile_item_id)
            except Inventory.DoesNotExist:
                return Response({"error": "Inventory not found for the given item ID"}, status=status.HTTP_404_NOT_FOUND)

            if not inventory.item.returnable:
                return Response({"error": "Item is non-returnable and cannot have copies"}, status=status.HTTP_400_BAD_REQUEST)

            data_item['inventory'] = inventory.id
            response_data.append(data_item)

        # Save each item copy to the database
        saved_copies = []
        for data_item in response_data:
            copy_serializer = self.get_serializer(data=data_item)
            if copy_serializer.is_valid():
                copy_instance = copy_serializer.save()
                saved_copies.append(copy_serializer.data)
                print(f"Saved ItemCopy with ID {copy_instance.id}")

        # Return a JSON response with the created item copies
        print("Saved Item Copies:", saved_copies)
        return Response({"item_copies": saved_copies}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def get_borrowed_and_lost_items_count(self, request):
        borrowed_items_count = ItemCopy.objects.filter(
            is_borrowed=True,
            condition__in=['Good', 'Slightly Damaged']
        ).count()

        lost_items = ItemCopy.objects.filter(
            is_borrowed=True,
            condition='Lost'
        )

        lost_items_count = lost_items.count()
        
        lost_serializer = ItemCopySerializer(lost_items, many=True)
        
        response_data = {
            'borrowed_items_count': borrowed_items_count,
            'lost_items_count': lost_items_count,
            'lost_items': lost_serializer.data
        }
        return Response(response_data)
    
    
class EditItemCopyStatusViewSet(viewsets.ModelViewSet):
    queryset = ItemCopy.objects.all()
    serializer_class = ItemCopySerializer

    def get_object(self):
        return get_object_or_404(ItemCopy, inventory__id=self.kwargs['inventory_id'], id=self.kwargs['item_id'])

class InventoryPagination(PageNumberPagination):
    page_size = 30
    page_size_query_param = 'page_size'
    max_page_size = 1000

class InventoryViewSet(viewsets.ModelViewSet):
    queryset = Inventory.objects.all()
    pagination_class = InventoryPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['item__name']

    def get_queryset(self):
        queryset = super().get_queryset().order_by('-quantity')
        search_query = self.request.query_params.get('search', None)
        category_filter = self.request.query_params.get('search_category', None)
        hidden_filter = self.request.query_params.get('hidden', None)
        if category_filter:
            # If category_filter is provided, filter items based on the category name
            queryset = queryset.filter(item__category__name__icontains=category_filter)
        elif search_query:
            # Apply search filter if search_query is provided
            queryset = queryset.filter(
                Q(item__name__icontains=search_query) |
                Q(item__category__name__icontains=search_query)
            )
            
        # Apply hidden filter
        if hidden_filter in ['true', 'True']:
            queryset = queryset.filter(is_hidden=True)
        elif hidden_filter in ['false', 'False']:
            queryset = queryset.filter(is_hidden=False)

        return queryset.order_by("id")

    #Different Serializer for Creation
    def get_serializer_class(self):
        if self.action == 'create':
            return InventoryCreateSerializer
        return InventorySerializer

    def create(self, request, *args, **kwargs):
        response_data = []

        for data_item in request.data:
            serializer = self.get_serializer(data=data_item)
            serializer.is_valid(raise_exception=True)

            item_id = data_item.get("item")
            quantity_to_add = serializer.validated_data.get("quantity", 0)

            inventory_instance, created = Inventory.objects.get_or_create(item_id=item_id, defaults={"quantity": 0})
            inventory_instance.quantity += quantity_to_add
            inventory_instance.save()

            response_data.append(InventorySerializer(inventory_instance).data)

        return Response(response_data, status=status.HTTP_201_CREATED)
    @action(detail=False, methods=['get'])
    def get_category_total_quantity(self, request):
        category_name = request.GET.get('category_name')
        include_hidden = request.GET.get('include_hidden', 'true').lower()  # Default to true if parameter is not provided or invalid

        queryset = Inventory.objects.filter(item__category__name__icontains=category_name)

        # Exclude hidden items if include_hidden parameter is set to false
        if include_hidden == 'false':
            queryset = queryset.filter(is_hidden=False)

        total_quantity = queryset.aggregate(Sum('quantity'))['quantity__sum']
        total_quantity = total_quantity if total_quantity is not None else 0
        return JsonResponse({'total_quantity': total_quantity})
        
    
            
            
class ExportMultipleTablesView(APIView):
    def get(self, request, *args, **kwargs):
        # Construct filename with the current date and time
        current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename = f'VirtualRorData_{current_datetime}.xlsx'
        
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        start_date = parse_date(start_date_str) if start_date_str else None
        end_date = parse_date(end_date_str) if end_date_str else None
        # Specify desktop path
        desktop_path = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')

        # Construct full file path with desktop path
        full_file_path = os.path.join(desktop_path, filename)

        # Create Excel writer object with the constructed full file path
        output = pd.ExcelWriter(full_file_path, engine='xlsxwriter')

        # Fetch data from the Inventory table with selected columns
        inventory_queryset = Inventory.objects.all().annotate(
            ItemID=F('id'),
            ItemName=F('item__name'),
            Type=F('item__returnable'),
            Quantity=F('quantity'),
            Reserved=F('reserved_quantity')
        ).values('ItemID', 'ItemName', 'Type', 'Quantity', 'Reserved')
        
        # Apply default ascending order by id
        inventory_queryset = inventory_queryset.order_by('ItemID')

        # Replace True and False with Borrowable and Consumable
        for item in inventory_queryset:
            item['Type'] = 'Borrowable' if item['Type'] else 'Consumable'

        # Convert inventory queryset to a DataFrame
        inventory_df = pd.DataFrame(list(inventory_queryset))

        # Write inventory DataFrame to the first sheet
        inventory_df.to_excel(output, sheet_name='Inventory', index=False)
        
        inventory_column_widths = {
            'ItemID': 15,
            'ItemName': 30,
            'Type': 15,
            'Quantity': 15,
            'Reserved': 15,
        }
        # Access the worksheet for inventory
        inventory_worksheet = output.sheets['Inventory']

        # Set the width for each column in inventory DataFrame
        for col_name, width in inventory_column_widths.items():
            column_index = inventory_df.columns.get_loc(col_name)
            inventory_worksheet.set_column(column_index, column_index, width)
        

        # Fetch data from the ItemCopy table with selected columns
        item_copy_queryset = ItemCopy.objects.all().annotate(
            ItemGroup=F('inventory__id'),
            ItemName=F('inventory__item__name'),
            ItemID=F('display_id'),
            ItemCategory=F('inventory__item__category__name'),  # Annotate the item category
            Condition=F('condition'),
            Borrowed=F('is_borrowed'),
            Reserved=F('is_reserved'),
        ).order_by('ItemGroup').values('ItemGroup', 'ItemName', 'ItemID', 'ItemCategory', 'Condition', 'Borrowed', 'Reserved')

        # Convert item copy queryset to a DataFrame
        item_copy_df = pd.DataFrame(list(item_copy_queryset))

        # Write item copy DataFrame to the second sheet
        item_copy_df.to_excel(output, sheet_name='ItemCopies', index=False)
        
        
        item_copy_column_widths = {
            'ItemGroup': 15,
            'ItemName': 30,
            'ItemID': 15,
            'ItemCategory': 20,
            'Condition': 15,
            'Borrowed': 15,
            'Reserved': 15,
        }   
        
        # Access the worksheet for item copies
        item_copy_worksheet = output.sheets['ItemCopies']

        # Set the width for each column in item copy DataFrame
        for col_name, width in item_copy_column_widths.items():
            column_index = item_copy_df.columns.get_loc(col_name)
            item_copy_worksheet.set_column(column_index, column_index, width)
                

        # Fetch data from the Transaction model
        transaction_queryset = Transaction.objects.all().annotate(
            TransactionID=F('id'),
            Type=F('transaction_type'),
            DateCreated=F('date_created'),
            ParticipantName=Concat(F('participant__first_name'), V(' '), F('participant__last_name')),  
            Items=Count('transaction_items'),
            Active=F('is_active'),
        ).values('TransactionID','Type','DateCreated','ParticipantName','Items','Active')
        
        if start_date and end_date:
            transaction_queryset = transaction_queryset.filter(date_created__range=(start_date, end_date))

        if transaction_queryset.exists():
            # Convert transaction queryset to a DataFrame
            transaction_df = pd.DataFrame(list(transaction_queryset))
            
            # Modify the DataFrame if necessary
            transaction_df['DateCreated'] = transaction_df['DateCreated'].apply(lambda x: x.replace(tzinfo=None))

            # Set column widths and other formatting for the Transactions sheet
            transaction_column_widths = {
                'TransactionID': 15,
                'DateCreated': 20,
                'ParticipantName': 20,
                # Add other column widths as needed
            }
            # Write transaction DataFrame to the third sheet
            transaction_df.to_excel(output, sheet_name='Transactions', index=False)
            worksheet = output.sheets['Transactions']
            for col_name, width in transaction_column_widths.items():
                column_index = transaction_df.columns.get_loc(col_name)
                worksheet.set_column(column_index, column_index, width)

        # Fetch data from the Inquiry model
        inquiry_queryset = Inquiry.objects.all().annotate(
            InquiryID=F('id'),
            InquirerName=Concat(F('inquirer__first_name'), V(' '), F('inquirer__last_name')),  
            InquiryType=F('inquiry_type'),
            Status=F('status'),
            DateCreated=F('date_created'),
            ReservedItems=Count('reserved_items'),
        ).values('InquiryID', 'InquirerName', 'InquiryType', 'Status', 'DateCreated', 'ReservedItems')
        
        if start_date and end_date:
            inquiry_queryset = inquiry_queryset.filter(date_created__range=(start_date, end_date))

        if inquiry_queryset.exists():
            # Convert inquiry queryset to a DataFrame
            inquiry_df = pd.DataFrame(list(inquiry_queryset))
            
            # Modify the DataFrame if necessary
            inquiry_df['DateCreated'] = inquiry_df['DateCreated'].apply(lambda x: x.replace(tzinfo=None))
            
            # Set column widths and other formatting for the Inquiry sheet
            inquiry_column_widths = {
                'InquiryID': 15,
                'InquirerName': 30,
                'InquiryType': 20,
                'Status': 15,
                'DateCreated': 20,
                'ReservedItems': 15,
            }
            # Write inquiry DataFrame to the fourth sheet
            inquiry_df.to_excel(output, sheet_name='Inquiry', index=False)
            inquiry_worksheet = output.sheets['Inquiry']
            for col_name, width in inquiry_column_widths.items():
                column_index = inquiry_df.columns.get_loc(col_name)
                inquiry_worksheet.set_column(column_index, column_index, width)
                 
        # Close the Excel writer object
        output.close()

         # Record export history
        export_history = ExportHistory.objects.create(filename=filename)
        
        # Load the workbook using openpyxl
        workbook = load_workbook(full_file_path)

        # Get the 'ItemCopies' sheet
        item_copies_sheet = workbook['ItemCopies']

        # Define red fill
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

        # Iterate over each row (excluding the header row)
        for row in item_copies_sheet.iter_rows(min_row=2, max_row=item_copies_sheet.max_row, min_col=6, max_col=7):
            for cell in row:
                # Check if the cell value is True and apply red fill
                if cell.value:
                    cell.fill = red_fill

        # Save the modified workbook
        workbook.save(full_file_path)

        # Open and read the generated Excel file
        with open(full_file_path, 'rb') as f:
            file_data = f.read()

        # Delete the workbook file
        os.remove(full_file_path)
        
        # Return the Excel file as a response
        response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    


# Assuming item_copy_id is the ID of the ItemCopy you want to find transactions for
def get_transactions_for_item_copy(item_copy_id):
    # Query TransactionItem model to filter transactions based on the item copy ID
    transactions = TransactionItem.objects.filter(item__id=item_copy_id).values_list('transaction__id', flat=True).distinct()
    return transactions
class ItemCopyTransactionsView(APIView):
    def get(self, request, item_copy_id):
        transactions = get_transactions_for_item_copy(item_copy_id)
        transactions = Transaction.objects.filter(id__in=transactions)
        transactions = transactions.order_by('-date_created')
        serializer = TransactionSerializer(transactions, many=True)
        return Response(serializer.data)
    
class ExportHistoryViewSet(viewsets.ModelViewSet):
    queryset = ExportHistory.objects.all()
    serializer_class = ExportHistorySerializer